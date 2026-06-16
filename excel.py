# faça um pip install pandas openpyxl psutil antes de rodar os scripts se preferir tambem utilize um venv

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def processar_planilha_telemetria(arquivo_entrada: str):
    """Lê planilha Excel de telemetria e gera relatório consolidado."""
    
    caminho = Path(arquivo_entrada)
    if not caminho.exists():
        caminho = Path("telemetria_recebida") / caminho.name
        if not caminho.exists():
            print(f"❌ Arquivo não encontrado: {arquivo_entrada}")
            return None
    
    df = pd.read_excel(caminho)
    
    # Padronizar nomes de colunas
    df.columns = [col.strip().lower().replace(" ", "_") for col in df.columns]
    
    # === Cálculo de Indicadores Principais ===
    indicadores = {
        'Data_Processamento': pd.Timestamp.now(),
        'Total_Registros': len(df),
        'Arquivo_Origem': caminho.name,
    }
    
    # Campos principais de telemetria
    if 'temp_obc_c' in df.columns:
        indicadores.update({
            'Temperatura_OBC_Media': round(df['temp_obc_c'].mean(), 2),
            'Temperatura_OBC_Max': round(df['temp_obc_c'].max(), 2),
            'Temperatura_OBC_Min': round(df['temp_obc_c'].min(), 2),
        })
    
    if 'soc_bateria_pct' in df.columns:
        indicadores.update({
            'SOC_Bateria_Media': round(df['soc_bateria_pct'].mean(), 2),
            'SOC_Bateria_Min': round(df['soc_bateria_pct'].min(), 2),
        })
    
    if 'em_eclipse' in df.columns:
        indicadores['Registros_Em_Eclipse'] = int(df['em_eclipse'].sum())
    
    # Contagem de estados anormais
    estado_cols = [col for col in df.columns if col.startswith('estado_')]
    for col in estado_cols:
        if col in df.columns:
            anomalias = len(df[df[col].str.contains('SAFE_MODE|DEGRADED|OFF|Anormal', na=False)])
            indicadores[f'{col.replace("estado_", "")}_Anomalias'] = anomalias
    
    df_indicadores = pd.DataFrame([indicadores])
    
    nome_saida = f"RELATORIO_TELEMETRIA_{Path(caminho).stem}_{pd.Timestamp.now():%Y%m%d_%H%M%S}.xlsx"
    
    with pd.ExcelWriter(nome_saida, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Dados_Brutos', index=False)
        df_indicadores.to_excel(writer, sheet_name='Indicadores', index=False)
    
    # Formatação
    try:
        wb = load_workbook(nome_saida)
        ws = wb['Indicadores']
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        wb.save(nome_saida)
    except:
        pass
    
    print(f"✅ Planilha processada com sucesso!")
    print(f"📁 Relatório salvo como: {nome_saida}")
    print(f"📊 Total de registros: {len(df):,}")
    return df_indicadores

if __name__ == "__main__":
    processar_planilha_telemetria("telemetria_satelite.xlsx")